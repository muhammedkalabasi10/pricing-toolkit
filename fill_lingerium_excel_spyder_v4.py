#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Lingerium ürün sayfalarından verileri çekip Trendyol Excel şablonuna yazar.

Özellikler:
- Tek URL ile çalışabilir
- .txt dosyasındaki URL listesini satır satır işleyebilir
- Yerel HTML dosyası ile çalışabilir
- Aynı barkod varsa mevcut satırı günceller
- Yeni barkod varsa sona ekler
- Çıktı dosyası sorulmaz; çalışılan klasöre otomatik kaydeder
- Stokta olmayan renk ve bedenleri tabloya ekler; stoklarını 0 yazar
- URL ile çalışırken her renk için önce butondaki gerçek linki dener; böylece o renge ait görselleri ve bedenleri daha doğru çeker

Komut satırı örnekleri:
  python fill_lingerium_excel_spyder_v9.py --url "https://..." --template "pijama-takimi.xlsx"
  python fill_lingerium_excel_spyder_v9.py --url-list "urunler.txt" --template "pijama-takimi.xlsx"
  python fill_lingerium_excel_spyder_v9.py --html "urun.html" --template "pijama-takimi.xlsx"
"""

import argparse
import json
import re
import sys
import unicodedata
from collections import Counter
from copy import copy
from pathlib import Path
from urllib.parse import parse_qs, urlencode, urljoin, urlparse, urlunparse

import openpyxl
import requests
from bs4 import BeautifulSoup


DEFAULT_OUTPUT = "pijama-takimi_doldurulmus.xlsx"
DEFAULT_SHEET = "Ürünlerinizi Burada Listeleyin"


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def normalize_tr(text: str) -> str:
    mapping = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU")
    text = (text or "").translate(mapping)
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def slugify(text: str) -> str:
    text = normalize_tr(text).lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def parse_price(text: str):
    text = clean_text(str(text or ""))
    text = (
        text.replace("₺", "")
        .replace("TL", "")
        .replace("TRY", "")
        .replace("Türk Lirası", "")
        .strip()
    )
    text = re.sub(r"[^\d,\.]", "", text)
    if not text:
        return None

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        left, right = text.rsplit(",", 1)
        if len(right) == 2:
            text = left.replace(".", "").replace(",", "") + "." + right
        else:
            text = text.replace(",", "")
    elif "." in text:
        left, right = text.rsplit(".", 1)
        if len(right) != 2:
            text = text.replace(".", "")

    try:
        return float(text)
    except ValueError:
        return None


def color_abbr(color: str) -> str:
    norm = re.sub(r"[^A-Z0-9]", "", normalize_tr(color).upper())
    return norm[:3]




def normalize_barcode_key(value) -> str:
    text = clean_text(str(value or "")).upper()
    return re.sub(r"\s+", "", text)


def dedupe_existing_barcodes(ws, headers: dict):
    barcode_col = headers.get("Barkod")
    if not barcode_col:
        return 0

    first_seen = {}
    rows_to_delete = []
    for row_idx in range(2, ws.max_row + 1):
        key = normalize_barcode_key(ws.cell(row_idx, barcode_col).value)
        if not key:
            continue
        if key in first_seen:
            rows_to_delete.append(row_idx)
        else:
            first_seen[key] = row_idx

    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx, 1)

    return len(rows_to_delete)

def fetch_html(url: str) -> str:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/126.0.0.0 Safari/537.36"
        )
    }
    response = requests.get(url, headers=headers, timeout=30)
    response.raise_for_status()
    return response.text


def detect_product_name(soup: BeautifulSoup) -> str:
    for tag_name in ["h1", "h2"]:
        tag = soup.find(tag_name)
        if tag:
            txt = clean_text(tag.get_text(" ", strip=True))
            if txt and "Yorumlar" not in txt:
                return txt

    text = soup.get_text("\n")
    match = re.search(r"ESPUAR\s+(.+?)\s+Ürün Kodu:", text, re.S)
    if match:
        return clean_text(match.group(1))
    return ""


def detect_brand(soup: BeautifulSoup, product_name: str) -> str:
    brand_box = soup.select_one(".brand-name")
    if brand_box:
        txt = clean_text(brand_box.get_text(" ", strip=True))
        if txt:
            return txt.title()

    if product_name:
        return clean_text(product_name.split()[0].title())
    return ""


def detect_product_code(soup: BeautifulSoup) -> str:
    box = soup.select_one(".categories-detail")
    if box:
        text = clean_text(box.get_text(" ", strip=True))
        match = re.search(r"Ürün Kodu:\s*(.+)", text)
        if match:
            return clean_text(match.group(1))

    text = soup.get_text("\n")
    match = re.search(r"Ürün Kodu:\s*([^\n]+)", text)
    return clean_text(match.group(1)) if match else ""


def detect_price(soup: BeautifulSoup):
    selectors = [
        ".product-detail-page-detail-price-box .sell-price",
        ".product-detail-page-detail-box .sell-price",
        ".price-main .sell-price",
        ".sell-price",
        "[data-price]",
        "[content][itemprop='price']",
    ]

    for selector in selectors:
        for tag in soup.select(selector):
            candidates = [
                tag.get("data-price"),
                tag.get("content"),
                clean_text(tag.get_text(" ", strip=True)),
            ]
            for candidate in candidates:
                price = parse_price(candidate)
                if price is not None and price > 0:
                    return price

    html = str(soup)
    regexes = [
        r'"price"\s*:\s*"([\d\.,]+)"',
        r'"price"\s*:\s*([\d\.,]+)',
        r'"salePrice"\s*:\s*"([\d\.,]+)"',
        r'₺\s*([\d\.,]+)',
    ]
    for pattern in regexes:
        match = re.search(pattern, html, re.I)
        if match:
            price = parse_price(match.group(1))
            if price is not None and price > 0:
                return price

    return None


def is_out_of_stock_variant_box(tag) -> bool:
    if not tag:
        return False

    classes = tag.get("class") or []
    if "variant-types-out-of-stock" in classes:
        return True

    for child in tag.children:
        if getattr(child, "name", None) == "div":
            return True
    return False


def extract_variant_option_url(box, page_url: str | None = None) -> str | None:
    if not box:
        return None

    candidate_values = []

    def add_candidate(value):
        value = clean_text(str(value or ""))
        if not value:
            return
        if value.startswith(("javascript:", "#")):
            return
        candidate_values.append(value)

    nodes = [box]
    nodes.extend(list(box.parents)[:4])
    nodes.extend(box.find_all(["a", "button", "div"], limit=5))

    for node in nodes:
        if not getattr(node, "attrs", None):
            continue

        for attr in ("href", "data-href", "data-url", "data-link", "data-product-url", "data-variant-url"):
            add_candidate(node.get(attr))

        onclick = node.get("onclick") or ""
        match = re.search(
            r"""(?:location\.href|window\.location(?:\.href)?|document\.location(?:\.href)?)\s*=\s*['"]([^'"]+)['"]""",
            onclick,
            re.I,
        )
        if match:
            add_candidate(match.group(1))

    for value in candidate_values:
        if page_url:
            return urljoin(page_url, value)
        return value

    return None


def parse_variant_sections(soup: BeautifulSoup, page_url: str | None = None) -> dict:
    sections = {}
    for section in soup.select("div.mb-4"):
        label = section.select_one(".variant-type.choce-variant-type")
        if not label:
            continue

        name = clean_text(label.get_text(" ", strip=True))
        options = []

        for box in section.select(".variant-types"):
            span = box.select_one(".variant-name")
            value = clean_text(span.get_text(" ", strip=True)) if span else ""
            if not value:
                continue

            classes = box.get("class") or []
            selected = "selected-circle" in classes or "border-black" in classes or "border-gray-900" in classes

            option = {
                "value": value,
                "out_of_stock": is_out_of_stock_variant_box(box),
                "selected": selected,
                "url": extract_variant_option_url(box, page_url=page_url),
            }

            if value not in [item["value"] for item in options]:
                options.append(option)

        if options:
            sections[name] = options

    return sections


def detect_selected_variant_value(soup: BeautifulSoup, variant_name: str) -> str:
    sections = parse_variant_sections(soup)
    for name, options in sections.items():
        if normalize_tr(name).lower() != normalize_tr(variant_name).lower():
            continue
        for option in options:
            if option.get("selected"):
                return option["value"]
    return ""


def get_variant_options(sections: dict, variant_name: str):
    target = normalize_tr(variant_name).lower()
    for name, options in sections.items():
        if normalize_tr(name).lower() == target:
            return options
    return []


def detect_variants(soup: BeautifulSoup, page_url: str | None = None) -> dict:
    variants = {}
    sections = parse_variant_sections(soup, page_url=page_url)
    for name, options in sections.items():
        values = []
        for option in options:
            if option["value"] not in values:
                values.append(option["value"])
        if values:
            variants[name] = values
    return variants


def detect_images(soup: BeautifulSoup):
    slider = soup.select_one("div.image-slider.product-detail-page-slider.relative")
    images = []
    if slider:
        for img in slider.find_all("img"):
            src = img.get("src") or ""
            if src.startswith("http") and src not in images:
                images.append(src)
    return images[:8]


def extract_next_data_page_specific(soup: BeautifulSoup) -> dict | None:
    script = soup.find("script", id="__NEXT_DATA__")
    if not script or not script.string:
        return None
    try:
        data = json.loads(script.string)
        return data.get("props", {}).get("pageProps", {}).get("pageSpecificData")
    except Exception:
        return None


def detect_myikas_image_root(html: str) -> str:
    matches = re.findall(
        r"https://cdn\.myikas\.com/images/([0-9a-f\-]{36})/[0-9a-f\-]{36}/image_\d+\.webp",
        html,
        re.I,
    )
    if not matches:
        return ""
    return Counter(matches).most_common(1)[0][0]


def build_myikas_image_url(image_root: str, image_id: str, size: int = 3840) -> str:
    if not image_root or not image_id:
        return ""
    return f"https://cdn.myikas.com/images/{image_root}/{image_id}/image_{size}.webp"


def extract_variant_rows_from_next_data(soup: BeautifulSoup, html: str):
    page_data = extract_next_data_page_specific(soup)
    if not page_data:
        return []

    variant_types = page_data.get("variantTypes") or []
    variants = page_data.get("variants") or []
    image_root = detect_myikas_image_root(html)
    if not image_root:
        return []

    variant_type_name_by_id = {}
    for item in variant_types:
        variant_type = item.get("variantType") or {}
        vt_id = variant_type.get("id")
        vt_name = clean_text(variant_type.get("name"))
        if vt_id and vt_name:
            variant_type_name_by_id[vt_id] = vt_name

    rows = []
    for variant in variants:
        if variant.get("deleted"):
            continue

        color = ""
        size = ""
        for vv in variant.get("variantValues") or []:
            vt_name = variant_type_name_by_id.get(vv.get("variantTypeId"), "")
            if normalize_tr(vt_name).lower() == "renk":
                color = clean_text(vv.get("name"))
            elif normalize_tr(vt_name).lower() == "beden":
                size = clean_text(vv.get("name"))

        if not color or not size:
            continue

        image_urls = []
        for image_item in variant.get("images") or []:
            image_id = ""
            if isinstance(image_item, dict):
                image_id = (
                    (image_item.get("image") or {}).get("id")
                    or image_item.get("imageId")
                    or ""
                )
            if not image_id:
                continue
            url = build_myikas_image_url(image_root, image_id)
            if url and url not in image_urls:
                image_urls.append(url)

        rows.append({
            "color": color,
            "size": size,
            "images": image_urls[:8],
            "stock": variant.get("stock"),
            "sku": clean_text(variant.get("sku")),
            "variant_id": variant.get("id"),
            "is_active": variant.get("isActive"),
        })

    return rows


def attach_color_details_from_variant_rows(product: dict):
    rows = product.get("variant_rows") or []
    if not rows:
        product["color_details"] = {color: {"sizes": product.get("sizes", []), "images": product.get("images", []), "url": None} for color in product.get("colors", [])}
        product.setdefault("image_warnings", [])
        return product

    color_details = {}
    for row in rows:
        color = row["color"]
        color_details.setdefault(color, {"sizes": [], "images": row.get("images", []), "url": None})
        if row["size"] not in color_details[color]["sizes"]:
            color_details[color]["sizes"].append(row["size"])
        if not color_details[color]["images"] and row.get("images"):
            color_details[color]["images"] = row["images"]

    product["color_details"] = color_details
    product.setdefault("image_warnings", [])
    return product


def detect_breadcrumb_category(soup: BeautifulSoup) -> str:
    crumbs = [clean_text(a.get_text(" ", strip=True)) for a in soup.select(".breadcrumbs .breadcrumb-item a")]
    return crumbs[-1] if crumbs else ""


def extract_page_description(soup: BeautifulSoup) -> str:
    text = soup.get_text("\n")
    match = re.search(r"Ürün\s+Açıklaması\s+(.+?)\s+Yorumlar", text, re.S)
    return clean_text(match.group(1)) if match else ""


def detect_model_number(product_name: str, product_code: str) -> str:
    match = re.search(r"\b(\d{2,5})\b", product_name or "")
    if match:
        return match.group(1)

    match = re.search(r"(\d{2,5})", product_code or "")
    return match.group(1) if match else ""


def build_description(product: dict) -> str:
    colors = ", ".join(product["colors"])
    sizes = ", ".join(product["sizes"])
    return (
        f'{product["product_name"]}, ev giyiminde şıklık ve rahatlığı bir arada sunan '
        f'saten dokulu bir pijama takımıdır. Biye detayları, önden düğmeli üst tasarımı '
        f've uzun kollu yapısıyla hem zarif bir görünüm hem de konforlu bir kullanım sağlar. '
        f'Yumuşak tuşesi sayesinde dinlenme, uyku ve ev içi kullanım için uygundur. '
        f'Ürün {colors} renk seçenekleri ve {sizes} beden alternatifleriyle sunulmaktadır. '
        f'Günlük homewear kullanımı için ideal, sade ve şık bir alt-üst takım tercihidir.'
    )


def parse_product(html: str, page_url: str | None = None) -> dict:
    soup = BeautifulSoup(html, "html.parser")

    product_name = detect_product_name(soup)
    brand = detect_brand(soup, product_name)
    product_code = detect_product_code(soup)
    price = detect_price(soup)
    variants = detect_variants(soup, page_url=page_url)
    variant_sections = parse_variant_sections(soup, page_url=page_url)

    colors = variants.get("Renk", [])
    sizes = variants.get("Beden", [])
    model_number = detect_model_number(product_name, product_code)
    model_code = f"{slugify(brand)}{model_number}" if model_number else slugify(brand)

    color_links = {}
    for option in get_variant_options(variant_sections, "Renk"):
        if option.get("url"):
            color_links[option["value"]] = option["url"]

    base_images = detect_images(soup)
    variant_rows = extract_variant_rows_from_next_data(soup, html)

    if colors:
        allowed_colors = {slugify(color) for color in colors}
        variant_rows = [row for row in variant_rows if slugify(row["color"]) in allowed_colors]
    if sizes:
        allowed_sizes = {clean_text(size).upper() for size in sizes}
        variant_rows = [row for row in variant_rows if clean_text(row["size"]).upper() in allowed_sizes]

    if variant_rows:
        ordered_colors = []
        for color in colors or [row["color"] for row in variant_rows]:
            if color not in ordered_colors and any(colors_match(color, row["color"]) for row in variant_rows):
                ordered_colors.append(color)
        if not ordered_colors:
            ordered_colors = list(dict.fromkeys(row["color"] for row in variant_rows))

        ordered_sizes = []
        for size in sizes or [row["size"] for row in variant_rows]:
            if size not in ordered_sizes and any(clean_text(size).upper() == clean_text(row["size"]).upper() for row in variant_rows):
                ordered_sizes.append(size)
        if not ordered_sizes:
            ordered_sizes = list(dict.fromkeys(row["size"] for row in variant_rows))

        colors = ordered_colors
        sizes = ordered_sizes
        if not base_images:
            base_images = next((row.get("images", []) for row in variant_rows if row.get("images")), [])

    product = {
        "brand": brand,
        "product_name": product_name,
        "product_code": product_code,
        "price": price,
        "colors": colors,
        "sizes": sizes,
        "category_name": detect_breadcrumb_category(soup),
        "images": base_images,
        "page_description": extract_page_description(soup),
        "model_number": model_number,
        "model_code": model_code,
        "color_links": color_links,
        "variant_rows": variant_rows,
        "variant_sections": variant_sections,
    }
    product["generated_description"] = build_description(product)
    attach_color_details_from_variant_rows(product)
    return product


def copy_row_style(ws, source_row: int, target_row: int, max_col: int):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        tgt = ws.cell(target_row, col)
        if src.has_style:
            tgt._style = copy(src._style)
        tgt.font = copy(src.font)
        tgt.fill = copy(src.fill)
        tgt.border = copy(src.border)
        tgt.alignment = copy(src.alignment)
        tgt.protection = copy(src.protection)
        tgt.number_format = src.number_format


def build_color_url(product_url: str, color: str, size: str | None = None) -> str:
    parsed = urlparse(product_url)
    query = parse_qs(parsed.query, keep_blank_values=True)
    query["Renk"] = [color]
    if size:
        query["Beden"] = [size]
    encoded_query = urlencode(query, doseq=True)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, encoded_query, parsed.fragment))


def colors_match(left: str, right: str) -> bool:
    return slugify(left) == slugify(right)


def ordered_unique(values):
    result = []
    seen = set()
    for value in values:
        key = clean_text(str(value or "")).upper()
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(clean_text(str(value)))
    return result


def normalize_stock_to_template(stock_value):
    if stock_value is None:
        return None
    try:
        return 0 if float(stock_value) <= 0 else 20
    except Exception:
        text = clean_text(str(stock_value)).lower()
        if text in {"0", "yok", "false", "none", "null"}:
            return 0
        return 20



def merge_variant_rows_with_html_options(
    color: str,
    color_variant_rows: list,
    size_options: list,
    color_images: list,
    fallback_sizes: list,
    base_images: list,
    force_zero_stock: bool = False,
    color_url: str | None = None,
):
    rows_by_size = {}
    for row in color_variant_rows:
        key = clean_text(row.get("size", "")).upper()
        if key:
            rows_by_size[key] = row

    ordered_sizes = []
    ordered_sizes.extend(option.get("value") for option in size_options)
    ordered_sizes.extend(row.get("size") for row in color_variant_rows)
    ordered_sizes.extend(fallback_sizes)
    ordered_sizes = ordered_unique(ordered_sizes)

    size_status_map = {
        clean_text(option.get("value", "")).upper(): bool(option.get("out_of_stock"))
        for option in size_options
        if clean_text(option.get("value", ""))
    }

    if not ordered_sizes:
        return []

    merged_rows = []
    for size in ordered_sizes:
        size_key = clean_text(size).upper()
        source_row = rows_by_size.get(size_key, {})

        # Stok kararında HTML'deki beden butonları kaynak olsun.
        # Buton stokluysa 20, stoksuzsa 0 yazılır.
        # HTML bilgisi yoksa ancak o zaman JSON stok bilgisine düşülür.
        if force_zero_stock:
            stock = 0
        elif size_key in size_status_map:
            stock = 0 if size_status_map.get(size_key, False) else 20
        else:
            stock = normalize_stock_to_template(source_row.get("stock"))
            if stock is None:
                stock = 20

        merged_rows.append({
            "color": color,
            "size": size,
            "images": source_row.get("images") or color_images or base_images,
            "stock": stock,
            "color_url": color_url,
            "sku": source_row.get("sku", ""),
            "variant_id": source_row.get("variant_id"),
            "is_active": source_row.get("is_active"),
        })

    return merged_rows


def enrich_product_with_color_pages(product: dict, product_url: str):
    base_sizes = product.get("sizes", [])
    base_images = product.get("images", [])
    base_color_links = product.get("color_links", {})
    base_sections = product.get("variant_sections") or {}
    base_color_options = get_variant_options(base_sections, "Renk")
    base_color_option_map = {slugify(opt.get("value")): opt for opt in base_color_options}

    existing_rows = product.get("variant_rows") or []
    existing_rows_by_color = {}
    for row in existing_rows:
        existing_rows_by_color.setdefault(slugify(row.get("color")), []).append(row)

    color_details = {}
    image_warnings = []
    new_variant_rows = []

    for color in product.get("colors", []):
        color_key = slugify(color)
        base_color_option = base_color_option_map.get(color_key, {})
        force_zero_stock = bool(base_color_option.get("out_of_stock"))

        candidate_urls = []
        explicit_url = base_color_links.get(color)
        fallback_url = build_color_url(product_url, color, base_sizes[0] if base_sizes else None)
        for url in (explicit_url, fallback_url):
            if url and url not in candidate_urls:
                candidate_urls.append(url)

        resolved_rows = []
        resolved_images = []
        resolved_url = explicit_url or fallback_url
        last_error = None

        for color_url in candidate_urls:
            try:
                color_html = fetch_html(color_url)
                color_soup = BeautifulSoup(color_html, "html.parser")
                selected_color = detect_selected_variant_value(color_soup, "Renk")
                if selected_color and not colors_match(selected_color, color):
                    continue

                color_sections = parse_variant_sections(color_soup, page_url=color_url)
                size_options = get_variant_options(color_sections, "Beden")
                color_variant_rows = extract_variant_rows_from_next_data(color_soup, color_html)
                color_variant_rows = [row for row in color_variant_rows if colors_match(row.get("color", ""), color)]
                color_images = next((row.get("images", []) for row in color_variant_rows if row.get("images")), [])
                color_images = color_images or detect_images(color_soup) or base_images

                merged_rows = merge_variant_rows_with_html_options(
                    color=color,
                    color_variant_rows=color_variant_rows,
                    size_options=size_options,
                    color_images=color_images,
                    fallback_sizes=base_sizes,
                    base_images=base_images,
                    force_zero_stock=force_zero_stock,
                    color_url=color_url,
                )

                if merged_rows:
                    resolved_rows = merged_rows
                    resolved_images = color_images
                    resolved_url = color_url
                    break
            except Exception as exc:
                last_error = str(exc)

        if not resolved_rows:
            fallback_variant_rows = existing_rows_by_color.get(color_key, [])
            fallback_size_options = get_variant_options(base_sections, "Beden") if base_sizes else []
            resolved_rows = merge_variant_rows_with_html_options(
                color=color,
                color_variant_rows=fallback_variant_rows,
                size_options=fallback_size_options,
                color_images=next((row.get("images", []) for row in fallback_variant_rows if row.get("images")), []) or base_images,
                fallback_sizes=base_sizes,
                base_images=base_images,
                force_zero_stock=force_zero_stock,
                color_url=resolved_url,
            )
            resolved_images = next((row.get("images", []) for row in resolved_rows if row.get("images")), []) or base_images

            if not resolved_rows:
                resolved_rows = [
                    {
                        "color": color,
                        "size": size,
                        "images": base_images,
                        "stock": 0 if force_zero_stock else 20,
                        "color_url": resolved_url,
                    }
                    for size in base_sizes
                ]

            warning = f"{product.get('model_code', '')} / {color} için renk sayfası tam çözümlenemedi; eldeki verilerle devam edildi."
            if last_error:
                warning += f" Hata: {last_error}"
            image_warnings.append(warning)

        new_variant_rows.extend(resolved_rows)
        color_details[color] = {
            "sizes": ordered_unique([row.get("size") for row in resolved_rows]),
            "images": resolved_images or base_images,
            "url": resolved_url,
        }

    if new_variant_rows:
        product["variant_rows"] = new_variant_rows
    product["color_details"] = color_details
    product["image_warnings"] = image_warnings
    return product


def attach_default_color_details(product: dict):
    return attach_color_details_from_variant_rows(product)


def build_variant_rows(product: dict):
    if product.get("variant_rows"):
        rows = []
        for row in product["variant_rows"]:
            stock = normalize_stock_to_template(row.get("stock"))
            if stock is None:
                stock = 20
            rows.append({
                "color": row["color"],
                "size": row["size"],
                "images": row.get("images") or product.get("images", []),
                "color_url": row.get("color_url"),
                "stock": stock,
            })
        return rows

    color_details = product.get("color_details") or {}
    rows = []

    if color_details:
        for color in product.get("colors", []):
            detail = color_details.get(color, {})
            sizes = detail.get("sizes") or product.get("sizes", [])
            for size in sizes:
                rows.append({
                    "color": color,
                    "size": size,
                    "images": detail.get("images") or product.get("images", []),
                    "color_url": detail.get("url"),
                    "stock": 20,
                })
        return rows

    colors = product["colors"]
    sizes = product["sizes"]
    return [{"color": color, "size": size, "images": product.get("images", []), "color_url": None, "stock": 20} for color in colors for size in sizes]


def find_last_data_row(ws, headers: dict) -> int:
    key_headers = ["Barkod", "Model Kodu", "Ürün Adı", "Stok Kodu", "Renk", "Beden"]
    key_cols = [headers[h] for h in key_headers if h in headers]
    if not key_cols:
        return 1

    for row_idx in range(ws.max_row, 1, -1):
        for col_idx in key_cols:
            value = ws.cell(row_idx, col_idx).value
            if value is not None and clean_text(str(value)):
                return row_idx
    return 1


def index_existing_barcodes(ws, headers: dict) -> dict:
    barcode_col = headers.get("Barkod")
    if not barcode_col:
        return {}

    barcode_rows = {}
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row_idx, barcode_col).value
        barcode = clean_text(str(value)) if value is not None else ""
        if barcode:
            barcode_rows[barcode] = row_idx
    return barcode_rows


def resolve_workbook_path(template_path: str, output_path: str) -> str:
    output_file = Path(output_path)
    if output_file.exists():
        return str(output_file)
    return template_path


def open_workbook_for_upsert(template_path: str, output_path: str):
    workbook_path = resolve_workbook_path(template_path, output_path)
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[DEFAULT_SHEET]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    removed_duplicate_rows = dedupe_existing_barcodes(ws, headers)
    if removed_duplicate_rows:
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    template_category = ws["D2"].value
    last_data_row = find_last_data_row(ws, headers)
    style_source_row = 2 if last_data_row < 2 else last_data_row
    next_empty_row = 2 if last_data_row < 2 else last_data_row + 1
    existing_barcodes = index_existing_barcodes(ws, headers)
    return {
        "wb": wb,
        "ws": ws,
        "headers": headers,
        "template_category": template_category,
        "style_source_row": style_source_row,
        "next_empty_row": next_empty_row,
        "existing_barcodes": existing_barcodes,
        "removed_duplicate_rows": removed_duplicate_rows,
    }


def upsert_product_to_sheet(session: dict, product: dict):
    ws = session["ws"]
    headers = session["headers"]
    template_category = session["template_category"]

    variant_rows = build_variant_rows(product)
    inserted_count = 0
    updated_count = 0

    for variant in variant_rows:
        color = variant["color"]
        size = variant["size"]
        variant_images = variant.get("images") or product.get("images", [])
        variant_stock = variant.get("stock")
        if variant_stock is None:
            variant_stock = 20
        abbr = color_abbr(color)
        model_code = product["model_code"]

        barcode = f"{model_code.upper()}-{abbr}-{size.upper()}"
        stock_code = f"STK-{model_code.upper()}-{abbr}-{size.upper()}"

        barcode_key = normalize_barcode_key(barcode)
        row_idx = session["existing_barcodes"].get(barcode_key)
        if row_idx:
            updated_count += 1
        else:
            row_idx = session["next_empty_row"]
            if row_idx != session["style_source_row"]:
                copy_row_style(ws, session["style_source_row"], row_idx, ws.max_column)
            session["existing_barcodes"][barcode_key] = row_idx
            session["next_empty_row"] += 1
            inserted_count += 1

        values = {
            "Barkod": barcode,
            "Model Kodu": model_code,
            "Marka": product["brand"],
            "Kategori": template_category,
            "Para Birimi": "TRY",
            "Ürün Adı": product["product_name"],
            "Ürün Açıklaması": product["generated_description"],
            "Piyasa Satış Fiyatı (KDV Dahil)": product["price"] + 500,
            "Trendyol'da Satılacak Fiyat (KDV Dahil)": product["price"],
            "Ürün Stok Adedi": variant_stock,
            "Stok Kodu": stock_code,
            "KDV Oranı": 10,
            "Beden": size,
            "Alt-Üst Takım": "Gömlek-Pantolon",
            "Kol Boyu": "Uzun",
            "Ürün Detayı": "Biye",
            "Ek Özellik": "Saten",
            "Ürün Tipi": "Düz",
            "Kumaş Tipi": "Saten",
            "Paket İçeriği": "2'li",
            "Boy": "Uzun",
            "Kol Tipi": "Uzun Kol",
            "Materyal": "Dokuma",
            "Yaka Tipi": "Gömlek Yaka",
            "Ortam": "Homewear",
            "Kapama Şekli": "Full Düğme Kapama",
            "Cinsiyet": "Kadın / Kız",
            "Kalıp": "Regular",
            "Desen": "Düz",
            "Sezon": "Tüm Sezonlar",
            "Parça Sayısı": "2 Parça",
            "Bel": "Normal Bel",
            "Persona": "Cool & Comfort",
            "Siluet": "Basic",
            "Yaş Grubu": "Yetişkin",
            "Renk": color,
            "Menşei": "TR",
            "Paça Tipi": "Düz Paça",
            "Web Color": color,
            "Koleksiyon": "Basic",
            "Paça Boyu": "Uzun",
        }

        for idx, img_url in enumerate(variant_images, start=1):
            values[f"Görsel {idx}"] = img_url

        for header, value in values.items():
            col_idx = headers.get(header)
            if col_idx:
                ws.cell(row_idx, col_idx).value = value

        if headers.get("Piyasa Satış Fiyatı (KDV Dahil)"):
            ws.cell(row_idx, headers["Piyasa Satış Fiyatı (KDV Dahil)"]).number_format = "#,##0.00"
        if headers.get("Trendyol'da Satılacak Fiyat (KDV Dahil)"):
            ws.cell(row_idx, headers["Trendyol'da Satılacak Fiyat (KDV Dahil)"]).number_format = "#,##0.00"

    return {
        "total": len(variant_rows),
        "inserted": inserted_count,
        "updated": updated_count,
    }


def save_session(session: dict, output_path: str):
    session["wb"].save(output_path)


def ask_nonempty(prompt_text: str) -> str:
    while True:
        value = input(prompt_text).strip()
        if value:
            return value
        print("Bu alan boş bırakılamaz.")


def ask_choice(prompt_text: str, valid_choices: dict, default_key: str) -> str:
    while True:
        raw = input(prompt_text).strip().lower()
        if not raw:
            return default_key
        if raw in valid_choices:
            return raw
        print("Geçerli bir seçim yapın:", ", ".join(valid_choices.keys()))


def read_url_list(txt_path: str):
    path = Path(txt_path)
    if not path.exists():
        raise FileNotFoundError(f"TXT dosyası bulunamadı: {txt_path}")

    urls = []
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        value = clean_text(line)
        if not value or value.startswith("#"):
            continue
        urls.append(value)

    if not urls:
        raise ValueError("TXT dosyasında işlenecek URL bulunamadı.")
    return urls


def collect_interactive_inputs():
    print("\nLingerium ürün aktarma programı")
    print("-" * 40)
    print("1) Tek ürün URL'si")
    print("2) .txt dosyasındaki URL listesi")
    print("3) Yerel HTML dosyası")

    source_type = ask_choice(
        "Kaynak türünü seçin [1/2/3] (varsayılan 1): ",
        {"1": "url", "2": "url_list", "3": "html"},
        default_key="1",
    )

    url = None
    url_list = None
    html_path = None

    if source_type == "1":
        url = ask_nonempty("Ürün URL'sini girin: ")
    elif source_type == "2":
        url_list = ask_nonempty("Her satırda bir URL olan .txt dosya yolunu girin: ")
    else:
        html_path = ask_nonempty("HTML dosya yolunu girin: ")

    template = ask_nonempty("Excel şablon yolunu girin: ")

    return {
        "url": url,
        "url_list": url_list,
        "html": html_path,
        "template": template,
    }


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--url", help="Tek ürün URL'si")
    parser.add_argument("--url-list", help="Her satırda bir URL olan TXT dosyası")
    parser.add_argument("--html", help="Yerel HTML dosyası")
    parser.add_argument("--template", help="Excel şablon yolu")
    parser.add_argument("--output", help="Çıktı Excel yolu")
    return parser.parse_args()


def validate_inputs(args_dict: dict):
    sources = [args_dict.get("url"), args_dict.get("url_list"), args_dict.get("html")]
    if sum(1 for s in sources if s) != 1:
        raise ValueError("Tek bir kaynak vermelisiniz: URL veya URL listesi veya HTML dosyası.")
    if not args_dict.get("template"):
        raise ValueError("Excel şablon yolu zorunludur.")


def ensure_price(product: dict, source_label: str = ""):
    if product.get("price") is not None:
        return

    print("\nUyarı: Ürün fiyatı sayfadan otomatik okunamadı.")
    if source_label:
        print("Kaynak:", source_label)

    while True:
        raw = input("Lütfen satış fiyatını girin (örn: 1530 veya 1530,00): ").strip()
        price = parse_price(raw)
        if price is not None and price > 0:
            product["price"] = price
            return
        print("Geçerli bir fiyat girin.")


def get_source_items(args_dict: dict):
    if args_dict.get("url"):
        return [{"kind": "url", "value": args_dict["url"]}]
    if args_dict.get("url_list"):
        seen = set()
        items = []
        for url in read_url_list(args_dict["url_list"]):
            key = clean_text(url)
            if key in seen:
                continue
            seen.add(key)
            items.append({"kind": "url", "value": url})
        return items
    return [{"kind": "html", "value": args_dict["html"]}]


def load_html_from_source(item: dict) -> str:
    if item["kind"] == "url":
        return fetch_html(item["value"])
    return Path(item["value"]).read_text(encoding="utf-8")


def process_sources(args_dict: dict):
    output_path = args_dict.get("output") or str((Path.cwd() / DEFAULT_OUTPUT).resolve())
    source_items = get_source_items(args_dict)
    session = open_workbook_for_upsert(args_dict["template"], output_path)

    overall = {
        "products_total": len(source_items),
        "products_ok": 0,
        "products_failed": 0,
        "variants_total": 0,
        "variants_inserted": 0,
        "variants_updated": 0,
        "failures": [],
        "removed_duplicate_rows": session.get("removed_duplicate_rows", 0),
    }

    for idx, item in enumerate(source_items, start=1):
        label = item["value"]
        print(f"\n[{idx}/{len(source_items)}] İşleniyor: {label}")
        try:
            html = load_html_from_source(item)
            product = parse_product(html, page_url=item["value"] if item["kind"] == "url" else None)
            if item["kind"] == "url":
                enrich_product_with_color_pages(product, item["value"])
            ensure_price(product, label)
            written = upsert_product_to_sheet(session, product)

            overall["products_ok"] += 1
            overall["variants_total"] += written["total"]
            overall["variants_inserted"] += written["inserted"]
            overall["variants_updated"] += written["updated"]

            print(f"   Tamam: {product['product_name']}")
            print(f"   Model kodu: {product['model_code']}")
            print(f"   Varyant: toplam={written['total']}, yeni={written['inserted']}, güncellenen={written['updated']}")
            for warning in product.get("image_warnings", []):
                print(f"   Uyarı: {warning}")
        except Exception as exc:
            overall["products_failed"] += 1
            overall["failures"].append({"source": label, "error": str(exc)})
            print(f"   Hata: {exc}")

    save_session(session, output_path)
    overall["output_path"] = output_path
    return overall


def main():
    cli_args = parse_args()
    has_cli_values = any([
        cli_args.url,
        cli_args.url_list,
        cli_args.html,
        cli_args.template,
        cli_args.output,
    ])

    if has_cli_values:
        args_dict = {
            "url": cli_args.url,
            "url_list": cli_args.url_list,
            "html": cli_args.html,
            "template": cli_args.template,
            "output": cli_args.output,
        }
    else:
        args_dict = collect_interactive_inputs()

    validate_inputs(args_dict)
    result = process_sources(args_dict)

    print("\nİşlem tamamlandı")
    print("Toplam ürün sayısı:", result["products_total"])
    print("Başarılı ürün sayısı:", result["products_ok"])
    print("Hatalı ürün sayısı:", result["products_failed"])
    print("Toplam işlenen varyant sayısı:", result["variants_total"])
    print("Yeni eklenen varyant sayısı:", result["variants_inserted"])
    print("Güncellenen varyant sayısı:", result["variants_updated"])
    print("Çıktı dosyası:", result["output_path"])
    print("Silinen eski tekrar satır sayısı:", result["removed_duplicate_rows"])
    print("Not: Program açılışta aynı barkodlu eski tekrar satırları temizler; sonra aynı barkod varsa satırı günceller, yeni barkodları sona ekler.")

    if result["failures"]:
        print("\nHata alınan ürünler:")
        for item in result["failures"]:
            print(f"- {item['source']} -> {item['error']}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print("\nHata oluştu:", exc)
        if sys.stdin.isatty():
            input("Çıkmak için Enter'a basın...")
        raise

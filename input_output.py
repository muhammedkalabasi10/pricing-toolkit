import re
import openpyxl
from pathlib import Path

# === DOSYA YOLLARI (her hafta sadece WEEKLY_PATH değişecek) ===
MASTER_PATH = Path("Master_Boxer_Fanila_Kulot_UYakaAtlet_Fanila0230_sadece_UrunBilgisi_alis_fiyatlari_erdemden_paketli_iskontolu10.xlsx")
WEEKLY_PATH = Path("input_offer.xlsx")
OUT_PATH    = Path("output_offer.xlsx")

SHIPPING_COST = 125
VAT_RATE = 0.10  # %10

def parse_turkish_number(val):
    """
    Şunları anlayacak şekilde sayıya çevirir:
    - 545,99
    - '545,99 TL'
    - '533,00 TL - 545,99 TL' (içindeki EN BÜYÜK sayıyı alır)
    - '-' / None -> None
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)

    s = str(val).strip()
    if s in {"", "-", "—"}:
        return None

    nums = re.findall(r"\d[\d\.\s]*[\,\.]?\d*", s)
    cleaned = []
    for n in nums:
        n = n.replace(" ", "")
        if "," in n and "." in n:
            n = n.replace(".", "").replace(",", ".")
        else:
            n = n.replace(",", ".")
        try:
            cleaned.append(float(n))
        except ValueError:
            pass
    return max(cleaned) if cleaned else None

def parse_percent_to_rate(val):
    """
    '18 %' -> 0.18
    '3,6 %' -> 0.036
    18 -> 0.18
    0.18 -> 0.18
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        x = float(val)
        return x / 100.0 if x > 1 else x

    s = str(val).strip()
    if s in {"", "-", "—"}:
        return None
    s = s.replace("%", "").replace("٪", "").strip()
    num = parse_turkish_number(s)
    if num is None:
        return None
    return num / 100.0 if num > 1 else num

def build_purchase_price_map(master_path: Path):
    """
    Master dosyadaki 'Sonuç' sayfasından:
    HB Ürün Id (SKU) -> Alış Fiyatı (İskontolu %10) map'i üretir.
    """
    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb["Sonuç"]

    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    sku_col = headers.get("HB Ürün Id")
    price_col = headers.get("Alış Fiyatı (İskontolu %10)") or headers.get("Alış Fiyatı")

    if not sku_col or not price_col:
        raise ValueError("Master dosyada 'HB Ürün Id' ve alış fiyatı sütunları bulunamadı.")

    mp = {}
    for r in range(2, ws.max_row + 1):
        sku = ws.cell(r, sku_col).value
        price = ws.cell(r, price_col).value
        if sku is None:
            continue
        sku_key = str(sku).strip()
        p = parse_turkish_number(price)
        if p is not None:
            mp[sku_key] = p
    return mp

def find_col(ws, predicate):
    """Başlık satırında (1. satır) predicate'e uyan sütunu bulur."""
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if predicate(str(h) if h is not None else ""):
            return c
    return None

def main():
    purchase_map = build_purchase_price_map(MASTER_PATH)

    wb = openpyxl.load_workbook(WEEKLY_PATH)
    ws = wb["Teklifler"] if "Teklifler" in wb.sheetnames else wb[wb.sheetnames[0]]

    sku_col = find_col(ws, lambda s: s.strip().lower() == "sku")
    kom_col = find_col(ws, lambda s: "komisyon" in s.lower() and "teklif" in s.lower() and "1" in s)
    max_col = find_col(ws, lambda s: ("katılabileceğiniz" in s.lower() or "katilabilece" in s.lower())
                                   and "maximum" in s.lower() and "1" in s)
    fiyatgir_col = find_col(ws, lambda s: s.strip().lower() == "fiyat gir")

    if not all([sku_col, kom_col, max_col, fiyatgir_col]):
        raise ValueError(
            f"Sütunlar bulunamadı. Bulunanlar: SKU={sku_col}, Komisyon1={kom_col}, MaxFiyat1={max_col}, FiyatGir={fiyatgir_col}"
        )

    # "Fiyat Gir" sütununun hemen sağına 2 yeni sütun ekle
    insert_at = fiyatgir_col + 1
    ws.insert_cols(insert_at, amount=2)
    ws.cell(1, insert_at).value = "Kar/Zarar (Alış iskonto)"
    ws.cell(1, insert_at + 1).value = "Kar/Zarar (Alış iskonto + KDV %10)"

    for r in range(2, ws.max_row + 1):
        sku = ws.cell(r, sku_col).value
        sku_key = str(sku).strip() if sku is not None else None

        max_price = parse_turkish_number(ws.cell(r, max_col).value)
        rate = parse_percent_to_rate(ws.cell(r, kom_col).value)

        # gerekli alanlar yoksa boş bırak
        if sku_key is None or max_price is None or rate is None:
            ws.cell(r, insert_at).value = None
            ws.cell(r, insert_at + 1).value = None
            continue

        purchase = purchase_map.get(sku_key)

        # Master'da yoksa SONUÇ yazma (boş bırak)
        if purchase is None:
            ws.cell(r, insert_at).value = None
            ws.cell(r, insert_at + 1).value = None
            continue

        net_after_commission = max_price * (1 - rate)

        profit = net_after_commission - SHIPPING_COST - purchase
        profit_vat = net_after_commission - SHIPPING_COST - (purchase * (1 + VAT_RATE))

        ws.cell(r, insert_at).value = round(profit, 2)
        ws.cell(r, insert_at + 1).value = round(profit_vat, 2)

    wb.save(OUT_PATH)
    print(f"Kaydedildi: {OUT_PATH}")

if __name__ == "__main__":
    main()
